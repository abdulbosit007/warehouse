# -*- coding: utf-8 -*-
"""
Generate 300 synthetic survey rows preserving original distributions.
Respects gender-stratified Likert differences (Kadın vs Erkek).
"""
import openpyxl
import sys
import random
import math
from collections import Counter

sys.stdout.reconfigure(encoding='utf-8')
random.seed(2026)

# ────────────────────────────────────────────────────────────────
# Load original data
# ────────────────────────────────────────────────────────────────
wb_in = openpyxl.load_workbook(r'C:\Users\abdul\OneDrive\Desktop\actual_data.xlsx')
ws_in = wb_in.active

TOTAL_ROWS = ws_in.max_row
TOTAL_COLS = ws_in.max_column

data = []
for r in range(1, TOTAL_ROWS + 1):
    row = [ws_in.cell(r, c).value for c in range(1, TOTAL_COLS + 1)]
    data.append(row)

# Column indices (0-based)
COL_TIMESTAMP = 0
COL_B = 1
LIKERT_COLS = list(range(2, 48))   # C..AV → indices 2..47
COL_GENDER = 48                     # AW
COL_AGE = 49                        # AX
COL_EDU = 50                        # AY
COL_MARITAL = 51                    # AZ
COL_EMPLOY = 52                     # BA
COL_INCOME = 53                     # BB

TARGET = 300

# ────────────────────────────────────────────────────────────────
# Helper: weighted random choice
# ────────────────────────────────────────────────────────────────
def weighted_choice(counter):
    """Pick a key from Counter proportionally."""
    population = list(counter.keys())
    weights = list(counter.values())
    return random.choices(population, weights=weights, k=1)[0]

def make_counts(dist_dict, n):
    """Turn a proportional counter into exactly n integer counts."""
    total = sum(dist_dict.values())
    probs = {k: v / total for k, v in dist_dict.items()}
    counts = {k: math.floor(p * n) for k, p in probs.items()}
    remainder = n - sum(counts.values())
    # distribute remainder to largest fractional parts
    fracs = sorted(probs.keys(), key=lambda k: (probs[k] * n) % 1, reverse=True)
    for k in fracs[:remainder]:
        counts[k] += 1
    return counts

# ────────────────────────────────────────────────────────────────
# Col B combos
# ────────────────────────────────────────────────────────────────
b_combo_counter = Counter()
for row in data:
    val = row[COL_B]
    if val:
        parts = tuple(sorted([p.strip() for p in str(val).split(',')]))
        b_combo_counter[parts] += 1

# ────────────────────────────────────────────────────────────────
# Gender distribution → exact counts
# ────────────────────────────────────────────────────────────────
gender_counter = Counter(row[COL_GENDER] for row in data if row[COL_GENDER])
gender_counts = make_counts(gender_counter, TARGET)
print("Gender counts:", gender_counts)

# ────────────────────────────────────────────────────────────────
# Per-column Likert distributions, split by gender
# ────────────────────────────────────────────────────────────────
likert_by_gender = {'Kadın': {}, 'Erkek': {}}
for g in ['Kadın', 'Erkek']:
    g_rows = [row for row in data if row[COL_GENDER] == g]
    for ci in LIKERT_COLS:
        vals = [row[ci] for row in g_rows if row[ci] is not None]
        likert_by_gender[g][ci] = Counter(vals)

# ────────────────────────────────────────────────────────────────
# Employment distributions (split by gender to avoid Ev hanımı for Erkek)
# ────────────────────────────────────────────────────────────────
employ_by_gender = {}
for g in ['Kadın', 'Erkek']:
    g_rows = [row for row in data if row[COL_GENDER] == g]
    employ_by_gender[g] = Counter(row[COL_EMPLOY] for row in g_rows if row[COL_EMPLOY])

print("\nEmployment by gender:")
for g, cnt in employ_by_gender.items():
    print(f"  {g}: {dict(cnt)}")

# ────────────────────────────────────────────────────────────────
# Age pool from original (resample with small jitter)
# ────────────────────────────────────────────────────────────────
age_pool = [int(row[COL_AGE]) for row in data if row[COL_AGE] is not None]
age_counter = Counter(age_pool)

def pick_age():
    """Sample an age from empirical distribution with minor random variation."""
    base = weighted_choice(age_counter)
    jitter = random.choice([-1, 0, 0, 0, 1])   # mostly exact, small jitter
    return max(14, min(70, base + jitter))

# ────────────────────────────────────────────────────────────────
# Other categorical distributions
# ────────────────────────────────────────────────────────────────
edu_counter = Counter(row[COL_EDU] for row in data if row[COL_EDU])
marital_counter = Counter(row[COL_MARITAL] for row in data if row[COL_MARITAL])
income_counter = Counter(row[COL_INCOME] for row in data if row[COL_INCOME])

# ────────────────────────────────────────────────────────────────
# Build list of (gender, row_index) to maintain exact gender counts
# ────────────────────────────────────────────────────────────────
gender_list = []
for g, cnt in gender_counts.items():
    gender_list.extend([g] * cnt)
random.shuffle(gender_list)

# ────────────────────────────────────────────────────────────────
# Generate timestamps (realistic date range Feb–May 2026)
# ────────────────────────────────────────────────────────────────
import datetime

start_dt = datetime.datetime(2026, 2, 27, 17, 0, 0)
end_dt   = datetime.datetime(2026, 5, 15, 23, 59, 59)
delta_sec = int((end_dt - start_dt).total_seconds())

ts_seconds = sorted(random.sample(range(delta_sec), TARGET))
timestamps = [(start_dt + datetime.timedelta(seconds=s)).strftime('%d.%m.%Y %H:%M:%S')
              for s in ts_seconds]

# ────────────────────────────────────────────────────────────────
# Generate 300 rows
# ────────────────────────────────────────────────────────────────
synthetic = []
for i in range(TARGET):
    gender = gender_list[i]
    row = [''] * TOTAL_COLS

    # Col A – timestamp
    row[COL_TIMESTAMP] = timestamps[i]

    # Col B – multi-select combo
    combo = weighted_choice(b_combo_counter)
    row[COL_B] = ', '.join(combo)

    # Col C..AV – Likert (gender-stratified per column)
    for ci in LIKERT_COLS:
        dist = likert_by_gender[gender].get(ci, Counter({5: 5, 4: 2, 3: 1}))
        row[ci] = weighted_choice(dist)

    # Col AW – gender
    row[COL_GENDER] = gender

    # Col AX – age
    row[COL_AGE] = pick_age()

    # Col AY – education
    row[COL_EDU] = weighted_choice(edu_counter)

    # Col AZ – marital
    row[COL_MARITAL] = weighted_choice(marital_counter)

    # Col BA – employment (gender-stratified)
    row[COL_EMPLOY] = weighted_choice(employ_by_gender[gender])

    # Col BB – income
    row[COL_INCOME] = weighted_choice(income_counter)

    synthetic.append(row)

# ────────────────────────────────────────────────────────────────
# Validation
# ────────────────────────────────────────────────────────────────
print(f"\nGenerated {len(synthetic)} rows")
gen_gender = Counter(r[COL_GENDER] for r in synthetic)
print(f"Gender: {dict(gen_gender)}")
gen_b = Counter(r[COL_B] for r in synthetic)
print(f"Col B: {dict(gen_b)}")
gen_edu = Counter(r[COL_EDU] for r in synthetic)
print(f"Edu: {dict(gen_edu)}")
gen_emp = Counter(r[COL_EMPLOY] for r in synthetic)
print(f"Employ: {dict(gen_emp)}")
gen_inc = Counter(r[COL_INCOME] for r in synthetic)
print(f"Income: {dict(gen_inc)}")

# Check no Erkek has Ev hanımı
erkek_ev = sum(1 for r in synthetic if r[COL_GENDER] == 'Erkek' and r[COL_EMPLOY] == 'Ev hanımı')
print(f"Erkek with 'Ev hanımı': {erkek_ev}  (should be 0)")

# Likert sanity
all_lik = [r[ci] for r in synthetic for ci in LIKERT_COLS if r[ci]]
lc = Counter(all_lik)
tot = sum(lc.values())
print(f"Pooled Likert proportions: { {k: round(v/tot, 3) for k,v in sorted(lc.items())} }")

# ────────────────────────────────────────────────────────────────
# Write to Excel
# ────────────────────────────────────────────────────────────────
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Synthetic_300"

for row_data in synthetic:
    ws_out.append(row_data)

out_path = r'C:\Users\abdul\OneDrive\Desktop\synthetic_300.xlsx'
wb_out.save(out_path)
print(f"\nSaved to: {out_path}")

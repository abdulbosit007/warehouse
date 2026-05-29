# -*- coding: utf-8 -*-
import openpyxl
import sys
import json
import random
import math
from collections import Counter

sys.stdout.reconfigure(encoding='utf-8')
random.seed(42)

wb = openpyxl.load_workbook(r'C:\Users\abdul\OneDrive\Desktop\actual_data.xlsx')
ws = wb.active

TOTAL_ROWS = ws.max_row   # 137 data rows (no header)
TOTAL_COLS = ws.max_column  # 54

# Read all data
data = []
for r in range(1, TOTAL_ROWS + 1):
    row = [ws.cell(r, c).value for c in range(1, TOTAL_COLS + 1)]
    data.append(row)

print(f"Loaded {len(data)} rows, {TOTAL_COLS} columns\n")

# Column index mapping (0-based)
COL_TIMESTAMP = 0   # A
COL_B = 1           # B - multi-select
# C..AV = indices 2..47 = Likert 1-5
LIKERT_COLS = list(range(2, 48))  # 46 columns (C=2 to AV=47)
COL_GENDER = 48     # AW - gender
COL_AGE = 49        # AX - age (free text)
COL_EDU = 50        # AY - education
COL_MARITAL = 51    # AZ - marital status
COL_EMPLOY = 52     # BA (col 53) - employment
COL_INCOME = 53     # BB (col 54) - income

# ── Analyze Column B (multi-select) ──────────────────────────────────────────
print("=== COL B (multi-select) ===")
b_option_counter = Counter()
b_combo_counter = Counter()
for row in data:
    val = row[COL_B]
    if val:
        parts = [p.strip() for p in str(val).split(',')]
        combo = tuple(sorted(parts))
        b_combo_counter[combo] += 1
        for p in parts:
            b_option_counter[p] += 1

print("Individual option frequencies:")
for opt, cnt in b_option_counter.most_common():
    print(f"  {cnt:3d}  {opt}")
print(f"\nUnique combos: {len(b_combo_counter)}")
print("Top combos:")
for combo, cnt in b_combo_counter.most_common(10):
    print(f"  {cnt:3d}  {', '.join(combo)}")

# ── Analyze Likert cols ───────────────────────────────────────────────────────
print("\n=== LIKERT COLS (C..AV) distribution ===")
likert_dist = {}
for ci in LIKERT_COLS:
    vals = [row[ci] for row in data if row[ci] is not None]
    counter = Counter(vals)
    col_letter = chr(65 + ci) if ci < 26 else 'A' + chr(65 + ci - 26)
    likert_dist[ci] = dict(sorted(counter.items()))
    print(f"  Col {col_letter}: {dict(sorted(counter.items()))}")

# Overall Likert value distribution (pooled)
all_likert = []
for ci in LIKERT_COLS:
    all_likert.extend([row[ci] for row in data if row[ci] is not None])
pool_counter = Counter(all_likert)
total_likert = sum(pool_counter.values())
print(f"\nPooled Likert: {dict(sorted(pool_counter.items()))}")
print(f"Proportions:   { {k: round(v/total_likert,3) for k,v in sorted(pool_counter.items())} }")

# ── Analyze categorical cols ──────────────────────────────────────────────────
cat_cols = {
    'GENDER': COL_GENDER,
    'EDU': COL_EDU,
    'MARITAL': COL_MARITAL,
    'EMPLOY': COL_EMPLOY,
    'INCOME': COL_INCOME,
}
print("\n=== CATEGORICAL COLS ===")
cat_dists = {}
for name, ci in cat_cols.items():
    vals = [row[ci] for row in data if row[ci] is not None]
    counter = Counter(vals)
    cat_dists[name] = counter
    print(f"\n{name} (col {ci+1}):")
    for k, v in counter.most_common():
        print(f"  {v:3d} ({v/len(vals)*100:5.1f}%)  {k}")

# ── Analyze Age (AX) ─────────────────────────────────────────────────────────
print("\n=== AGE (AX) ===")
ages = [row[COL_AGE] for row in data if row[COL_AGE] is not None]
ages_int = []
for a in ages:
    try:
        ages_int.append(int(a))
    except:
        print(f"  Non-numeric age: {a}")
ages_int.sort()
print(f"  Count: {len(ages_int)}")
print(f"  Min: {min(ages_int)}, Max: {max(ages_int)}")
print(f"  Mean: {sum(ages_int)/len(ages_int):.1f}")
age_counter = Counter(ages_int)
print(f"  Distribution: {dict(sorted(age_counter.items()))}")

# ── Analyze gender-stratified Likert ─────────────────────────────────────────
print("\n=== GENDER BREAKDOWN ===")
for g in ['Kadın', 'Erkek']:
    g_rows = [row for row in data if row[COL_GENDER] == g]
    print(f"\n{g}: {len(g_rows)} rows")
    all_vals = []
    for ci in LIKERT_COLS:
        all_vals.extend([row[ci] for row in g_rows if row[ci] is not None])
    gc = Counter(all_vals)
    tot = sum(gc.values())
    print(f"  Pooled Likert: { {k: round(v/tot,3) for k,v in sorted(gc.items())} }")

print("\n=== DONE ===")
